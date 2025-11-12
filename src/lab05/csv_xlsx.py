import csv
import os
import openpyxl

def validate_extension(path, ext):
    if not path.lower().endswith(ext.lower()):
        raise ValueError(f"Неверный формат: {path}. Ожидается {ext}")

def csv_to_xlsx(csv_path, xlsx_path):
    validate_extension(csv_path, '.csv')
    validate_extension(xlsx_path, '.xlsx')
    if not os.path.exists(csv_path): raise FileNotFoundError(f"CSV не найден: {csv_path}")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        data = list(csv.reader(f))
        if not data: raise ValueError("CSV файл пуст")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    wb.save(xlsx_path)
    print(f"CSV -> XLSX: {xlsx_path}")

def xlsx_to_csv(xlsx_path, csv_path):
    validate_extension(xlsx_path, '.xlsx')
    validate_extension(csv_path, '.csv')
    if not os.path.exists(xlsx_path): raise FileNotFoundError(f"XLSX не найден: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    if ws.max_row == 0 or ws.max_column == 0: raise ValueError("XLSX файл пуст")
    
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows([cell if cell is not None else '' for cell in row] for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row))
    
    print(f"XLSX -> CSV: {csv_path}")